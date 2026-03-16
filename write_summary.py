import sys, json
from openpyxl import load_workbook, Workbook

# Argumente:
# 1 = xlsxPath
# 2 = JSON mit topCompanies
# 3 = JSON mit topSkills

xlsx_path = sys.argv[1]
top_companies = json.loads(sys.argv[2])
top_skills = json.loads(sys.argv[3])

wb = load_workbook(xlsx_path)

# Sheet vorbereiten
if "Summary" in wb.sheetnames:
    ws = wb["Summary"]
    # bestehende Inhalte löschen
    wb.remove(ws)

ws = wb.create_sheet("Summary")

# Header
ws["A1"] = "Metrik"
ws["B1"] = "Wert"

row = 2

# Top Firmen
for i, entry in enumerate(top_companies, start=1):
    ws[f"A{row}"] = f"Top Firma {i}"
    ws[f"B{row}"] = f"{entry['name']} ({entry['count']})"
    row += 1

# Leerzeile
row += 1

# Top Kenntnisse
for i, entry in enumerate(top_skills, start=1):
    ws[f"A{row}"] = f"Top Kenntnis {i}"
    ws[f"B{row}"] = f"{entry['name']} ({entry['count']})"
    row += 1

wb.save(xlsx_path)

print(json.dumps({
    "status": "ok",
    "written_to": xlsx_path
}))