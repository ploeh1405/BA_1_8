import sys, json
from openpyxl import load_workbook

path = sys.argv[1]
wb = load_workbook(path, data_only=True)

if "Berichte" not in wb.sheetnames:
    raise SystemExit(f'Sheet "Berichte" nicht gefunden. Vorhanden: {wb.sheetnames}')

ws = wb["Berichte"]

headers = [c.value for c in ws[1]]
headers = [h if h is not None else "" for h in headers]

rows = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    obj = {}
    for h, v in zip(headers, row):
        obj[str(h)] = "" if v is None else v
    obj["__rowNumber"] = i
    obj["__xlsxPath"] = path
    rows.append(obj)

print(json.dumps(rows, ensure_ascii=False))